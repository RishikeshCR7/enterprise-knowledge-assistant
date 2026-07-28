import io
import logging
from typing import Tuple, Dict, Any
import pdfplumber
import docx
import openpyxl

logger = logging.getLogger(__name__)


class DocumentParser:
    @staticmethod
    def parse_pdf(file_bytes: bytes) -> Tuple[str, Dict[str, Any]]:
        """
        Extracts text page by page from PDF using pdfplumber.
        """
        extracted_pages = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            total_pages = len(pdf.pages)
            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text()
                if text:
                    extracted_pages.append(text)
        
        full_text = "\n\n".join(extracted_pages)
        metadata = {
            "page_count": total_pages,
            "char_count": len(full_text)
        }
        return full_text, metadata

    @staticmethod
    def parse_docx(file_bytes: bytes) -> Tuple[str, Dict[str, Any]]:
        """
        Extracts text paragraph by paragraph from DOCX using python-docx.
        """
        doc = docx.Document(io.BytesIO(file_bytes))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        full_text = "\n\n".join(paragraphs)
        metadata = {
            "paragraph_count": len(paragraphs),
            "char_count": len(full_text)
        }
        return full_text, metadata

    @staticmethod
    def parse_xlsx(file_bytes: bytes) -> Tuple[str, Dict[str, Any]]:
        """
        Extracts text sheet by sheet from XLSX using openpyxl.
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet_texts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_rows = []
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join([str(val) for val in row if val is not None])
                if row_str.strip():
                    sheet_rows.append(row_str)
            if sheet_rows:
                sheet_texts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_rows))
        
        full_text = "\n\n".join(sheet_texts)
        metadata = {
            "sheet_count": len(wb.sheetnames),
            "char_count": len(full_text)
        }
        return full_text, metadata

    @classmethod
    def parse_file(cls, filename: str, file_bytes: bytes) -> Tuple[str, Dict[str, Any]]:
        ext = filename.split(".")[-1].lower()
        if ext == "pdf":
            return cls.parse_pdf(file_bytes)
        elif ext in ["docx", "doc"]:
            return cls.parse_docx(file_bytes)
        elif ext in ["xlsx", "xls"]:
            return cls.parse_xlsx(file_bytes)
        elif ext in ["txt", "md", "csv"]:
            text = file_bytes.decode("utf-8", errors="ignore")
            return text, {"char_count": len(text)}
        else:
            raise ValueError(f"Unsupported file format extension: {ext}")
